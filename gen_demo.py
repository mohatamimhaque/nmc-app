import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'committee_members'

headers = ['sub_committee','name','role','designation','department','email','phone','facebook','linkedin','photo_filename','is_visible']

# Style header row
header_fill = PatternFill(start_color='3B5BDB', end_color='3B5BDB', fill_type='solid')
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

# sub_committee values: Adviser, Core Committee, Technical, Logistics, Design, Media, Registration
data = [
  # Advisers
  ['Adviser',          'Prof. Dr. Md. Abul Kashem Mia',  'Chief Adviser',      'Professor',              'Dept. of CSE, DUET',              'kashem@duet.ac.bd',   '01711-000001', 'fb.com/kashem',    'linkedin.com/in/kashem',    'kashem.jpg',    True ],
  ['Adviser',          'Dr. Md. Shahidul Islam',          'Adviser',            'Associate Professor',    'Dept. of Mathematics, DUET',      'shahid@duet.ac.bd',   '01711-000002', 'fb.com/shahidul',  'linkedin.com/in/shahidul',  'shahidul.jpg',  True ],
  ['Adviser',          'Dr. Nasrin Akter',                'Adviser',            'Assistant Professor',    'Dept. of Mathematics, DUET',      'nasrin@duet.ac.bd',   '01711-000003', 'fb.com/nasrin',    'linkedin.com/in/nasrin',    'nasrin.jpg',    True ],

  # Core Committee
  ['Core Committee',   'Md. Imtiaz Hossain',              'Convener',           'B.Sc. 4th Year',         'Dept. of EEE',                    'imtiaz@nmc.bd',       '01812-100001', 'fb.com/imtiaz',    'linkedin.com/in/imtiaz',    'imtiaz.jpg',    True ],
  ['Core Committee',   'Sumaiya Binte Zahir',             'Joint Convener',     'B.Sc. 4th Year',         'Dept. of CSE',                    'sumaiya@nmc.bd',      '01812-100002', 'fb.com/sumaiya',   'linkedin.com/in/sumaiya',   'sumaiya.jpg',   True ],
  ['Core Committee',   'Rashed Karim',                    'General Secretary',  'B.Sc. 3rd Year',         'Dept. of Mathematics',            'rashed@nmc.bd',       '01812-100003', 'fb.com/rashed',    'linkedin.com/in/rashed',    'rashed.jpg',    True ],
  ['Core Committee',   'Lamia Sultana',                   'Treasurer',          'B.Sc. 3rd Year',         'Dept. of Civil Engineering',      'lamia@nmc.bd',        '01812-100004', 'fb.com/lamia',     'linkedin.com/in/lamia',     'lamia.jpg',     True ],

  # Technical
  ['Technical',        'Md. Tanvir Hossain',              'Technical Lead',     'B.Sc. 4th Year',         'Dept. of EEE',                    'tanvir@nmc.bd',       '01913-200001', 'fb.com/tanvir',    'linkedin.com/in/tanvir',    'tanvir.jpg',    True ],
  ['Technical',        'Asif Mahmud',                     'Web Developer',      'B.Sc. 3rd Year',         'Dept. of CSE',                    'asif@nmc.bd',         '01913-200002', 'fb.com/asif',      'linkedin.com/in/asif',      'asif.jpg',      True ],
  ['Technical',        'Priya Datta',                     'IT Support',         'B.Sc. 2nd Year',         'Dept. of CSE',                    'priya@nmc.bd',        '01913-200003', 'fb.com/priya',     '',                          '',              True ],

  # Logistics
  ['Logistics',        'Sabbir Ahmed',                    'Logistics Manager',  'B.Sc. 3rd Year',         'Dept. of Civil Engineering',      'sabbir@nmc.bd',       '01715-300001', 'fb.com/sabbir',    'linkedin.com/in/sabbir',    'sabbir.jpg',    True ],
  ['Logistics',        'Mehedi Hassan',                   'Venue Coordinator',  'B.Sc. 2nd Year',         'Dept. of ME',                     'mehedi@nmc.bd',       '01715-300002', 'fb.com/mehedi',    '',                          'mehedi.jpg',    True ],
  ['Logistics',        'Farhan Kabir',                    'Transport Head',     'B.Sc. 3rd Year',         'Dept. of ME',                     'farhan@nmc.bd',       '01715-300003', 'fb.com/farhan',    'linkedin.com/in/farhan',    '',              False],

  # Design
  ['Design',           'Fahmida Akter',                   'Design Head',        'B.Sc. 2nd Year',         'Dept. of Architecture',           'fahmida@nmc.bd',      '01614-400001', 'fb.com/fahmida',   'linkedin.com/in/fahmida',   'fahmida.jpg',   True ],
  ['Design',           'Nadia Islam',                     'Graphic Designer',   'B.Sc. 3rd Year',         'Dept. of Architecture',           'nadia@nmc.bd',        '01614-400002', 'fb.com/nadia',     '',                          'nadia.jpg',     True ],

  # Media
  ['Media',            'Tasnim Sultana',                  'Media Coordinator',  'M.Sc. Student',          'Dept. of Mathematics',            'tasnim@nmc.bd',       '01816-500001', 'fb.com/tasnim',    'linkedin.com/in/tasnim',    'tasnim.jpg',    True ],
  ['Media',            'Arif Hossain',                    'Photographer',       'B.Sc. 4th Year',         'Dept. of EEE',                    'arif@nmc.bd',         '01816-500002', 'fb.com/arif',      '',                          'arif.jpg',      True ],
  ['Media',            'Shirin Akter',                    'Videographer',       'B.Sc. 3rd Year',         'Dept. of CSE',                    'shirin@nmc.bd',       '01816-500003', 'fb.com/shirin',    '',                          '',              True ],

  # Registration
  ['Registration',     'Rakibul Hasan',                   'Registration Head',  'B.Sc. 4th Year',         'Dept. of ME',                     'rakib@nmc.bd',        '01917-600001', 'fb.com/rakibul',   'linkedin.com/in/rakibul',   'rakibul.jpg',   True ],
  ['Registration',     'Nusrat Jahan',                    'Data Entry',         'B.Sc. 3rd Year',         'Dept. of CSE',                    'nusrat@nmc.bd',       '01917-600002', 'fb.com/nusrat',    'linkedin.com/in/nusrat',    'nusrat.jpg',    True ],
  ['Registration',     'Sumaia Khanam',                   'Volunteer',          'B.Sc. 1st Year',         'Dept. of Mathematics',            'sumaia@nmc.bd',       '01917-600003', 'fb.com/sumaia',    '',                          '',              False],
]

# Alternate row color
alt_fill = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid')
for row_idx, row in enumerate(data, 2):
    ws.append(row)
    if row_idx % 2 == 0:
        for cell in ws[row_idx]:
            cell.fill = alt_fill

# Auto column width
for col in ws.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    ws.column_dimensions[col[0].column_letter].width = max_len + 4

out = r'd:/nmc/nmc-app/committee_demo.xlsx'
wb.save(out)
print(f'Done — {len(data)} rows saved to {out}')
